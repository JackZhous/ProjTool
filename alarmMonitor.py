#!/bin/bash
#encoding=utf-8

# author: jackzhous
# 描述：和易充后台告警监控脚本
import sys
from openpyxl import *
import os
from charger import EasyCharger

imgLabel = {}

class AMonitor:
	def __init__(self, charger):
		self.charger = charger

	def initXls(self):
		self.book = Workbook()
		sheet = self.book.create_sheet('alarmManager', 0)
		sheet.append(['告警事件', '告警类型', '设备编码', '告警级别', '所属电站', '告警发生时间', '告警解决时间'])
		

	# 查询某个端口警告
	# time格式 2018-2-4~20182-5,从~分割
	def searchAlarmWithPort(self, port, time):
		self.initXls()
		timeArrays = time.split('~')
		result = self.getAlarm(port, timeArrays[0], timeArrays[1])
		return result		
	
	def searchAlarmWithStation(self, stationName, time):
		self.initXls()
		timeArrays = time.split('~')
		statData = self.searchStationPorts(stationName)
		self.saveDeviceToExcel(stationName, statData)
		ret = 0
		for key in statData:
                      	result = self. getAlarm(key, timeArrays[0], timeArrays[1])
			if result == 1:
				ret = 1
		return ret

	def searchAllStation(self, time):
		print time
                station = {"currentPage": 1, "itemsPerPage": 50000}
		if time is not None and time != '':
			time = time.split('~')
			station['startTime'] = time[0]
        	        station['endTime'] = time[1]
                result = self.charger.getAlarm(station)
		if result != 'false':
			return result['list']
		return 'false'

	def main(self):
		initUI()
	#	self.charger.login('1221')
	#	init(self)
	#	station = {"deviceImei": "5149013215964394",
	#		"startTime": "2018-03-06",
	#		"endTime": "2018-03-27",
	#		"currentPage": 1,
	#		"itemsPerPage": 10}
	#	saveAllStationNameToFile(self.charger)
	#	statData = searchStationPorts(self.charger, u'朗新科技')
	#	startTime = '2018-03-06'
	#	endTime = '2018-03-27'
	#	saveDeviceToExcel('aassa', statData)
	#	for key in statData: 
	#		getAlarm(self.charger, key, startTime, endTime)

	def getAlarm(self, deviceCode, startTime, endTime):
		station = {"currentPage": 1, "itemsPerPage": 10}
		station['deviceImei'] = deviceCode
		station['startTime'] = startTime
		station['endTime'] = endTime
		result = self.charger.getAlarm(station)
		station['itemsPerPage'] = result['total']
		result = self.charger.getAlarm(station)
		sheet = self.book.get_sheet_by_name('alarmManager')
		if result != 'false':
			return self.saveAlarmToExcel(sheet, result)
		return result



	# 入参data格式
	#{
	#	"code": 1,
	#	"message": "查询成功",
	#	"total": 5,
	#	"list": [{
	#		"firstDate": "2018-03-25 15:54:17",
	#		"remark": "",
	#		"handlePerson": "",
	#		"stationId": 253,
	#		"alarmNum": 1,
	#		"alarmId": "2018032515541748522402032",
	#		"alarmEventId": 5,
	#		"alarmRemark": "设备5149013215964394网络问题",
	#		"alarmStatus": 0,
	#		"newDate": "2018-03-25 15:54:17",
	#		"alarmEvent": "网络问题",
	#		"alarmLevel": 2,
	#		"deviceImei": "5149013215964394",
	#		"handleDate": "",
	#		"alarmType": 2,
	#		"deviceStation": "河南省洛阳市洛龙区太康东路中国移动洛阳呼叫基地新园区"
	#}
	def saveAlarmToExcel(self, sheet, stationData):
		list = stationData["list"]
		if len(list) <= 0:
			return 0
		for item in list:
			data = [item['alarmEvent'], item['alarmRemark'], item['deviceImei'], item['alarmLevel'], item['deviceStation'], item['firstDate'] + '\n' + item['newDate'], item['handleDate']]
			sheet.append(data)
		return 1

	def saveDeviceToExcel(self, staname, port):
		sheet = self.book.create_sheet('DeviceManger', 1)
		sheet.append(['充电桩名称', '端口号'])
		for key in port:
			sheet.append([staname, key])

	def saveAllStationNameToFile(self, stationFile):
		stations = self.charger.listAllProvinceData()
		if stations == 'false':
			return
		stationFile = open(stationFile, 'w')
		stationFile.write('---这是一个整合所有充电桩站点名字的文件，可以以站点名字为单位搜索每个站点的告警---\n')
		index = 0
		for item in stations:
			statname = item['staname'] + '  '
			if item['blug'] == 0 or  u'长虹' in statname or u'测试' in statname or u'演示' in statname:
				continue
			index = index + 1
			if index >= 5:
				index = 0
				statname = statname + '\n'
			stationFile.write(statname)
		stationFile.close()
		

	def searchStationPorts(self, stationName):
		if stationName is None:
			return None
		portsData = {}
		station = self.charger.searchStation(stationName)
		for item in station:
			if item['staname'] == stationName:
				stationPorts = self.charger.findPortPerStation(item['staid'], item['blug'])
				stationGateway = self.charger.findGateway(item['staid'])
				return self.filterStationPorts(stationPorts, stationGateway)
		return None

	#过滤电站端口，主要是相同_0和_1为一个
	def filterStationPorts(self, stationPorts, getway):
		data = {}
		for item in stationPorts:
			port = str(item['pgstanum'])
			data[port] = '0'

		for gate in getway:
			data[gate['devimeiStr']] = '0'
		return data	

	def saveFile(self, filename):
		self.book.save(filename)
		self.book.close()
	
if __name__ == '__main__':
	main()
	self.book.save('xx.xlsx')
