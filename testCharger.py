#!/bin/bash
#encoding=utf-8

# author: jackzhous
# 描述：网站爬下来的数据分析处理归类，保存于文档
import sys
from charger import EasyCharger 
from openpyxl import *
import datetime
from compare_excel import CompareEXCEL

import xlrd
class ChargerManager:
	
	def __init__(self, charger):
		self.charger = charger
	
	#获取所有的项目
	def getAllProject(self, xlsfile):
		self.charger = EasyCharger(xlsfile)
		self.charger.login()
		all_project = self.charger.listAllProvinceData()
		province = self.charger.getProvinceInfo()
		if all_project == 'false':
			return
		self.saveAllProject(xlsfile, all_project, province)

	def saveAllProject(self, xlsfile, all_project, province):
		book = Workbook()
		sheet0 = book.create_sheet(u'项目总览信息表', 0)
		sheet0.append(['项目名称', '项目位置', '省份'])
		for item in all_project:
			if u'测试' in item['staname'] or u'长虹' in item['staaddress']:
				continue
			if item['blug'] == 0:
				continue
			provinceId = int(item['province'])
			provinceItem = province[provinceId]
			data = [item['staname'], item['staaddress'], provinceItem['name']]
			sheet0.append(data)
		book.save(xlsfile)


	def getStationConfig(self, filename):
		province = self.charger.getProvinceInfo()
                result = self.charger.listAllProvinceData()
		if result == "false" or province == 'false':
                        return "false"
		allUnormalData = []
                for item in result:
                        if item["blug"] == 0 or item['staaddress'] == u'长虹':
                                continue
                        if u'测试' in item['staname']:
                                continue
                        provinceId = int(item['province'])
                        provinceItem = province[provinceId]
                        result1 = self.charger.getStationConfig(item['configid'])   #每个充电站的结果
                        if result1['code'] == 1:
				ret = result1['detail']
				if ret is None or ret == '':
					 data = {'staname':item['staname'], 'over_load_electric':'default', 'no_load_time': 'default', 'over_load_time':'default', 'full_electric':'default', 'under_load_electric':'default', 'heart_time':'default'}
				else:
					data = {'staname':item['staname'], 'over_load_electric':ret['over_load_electric'], 'no_load_time': ret['no_load_time'], 'over_load_time':ret['over_load_time'], 'full_electric':ret['full_electric'], 'under_load_electric':ret['under_load_electric']}
				allUnormalData.append(data)
                return self.saveAllStaConfData(allUnormalData, filename)


# 大数据结构展示
#
#	    ---- ‘staname’: {'proj_name', 'province', 'all_port', 'lost_port', 'error_port', 'new_lost_port'}
# all_proj{}
#
#
	#获取所有异常数据
	def getLostDataFromNetwork(self, oldFileData, filename, code, time):
		print("---start---")
		timeArrays = time.split('~')	
		province = self.charger.getProvinceInfo()
		result = self.charger.listAllProvinceData()
		if result == "false" or province == 'false':
			return "false"
		allUnormalData = []
		all_proj = {}
		all_order = []
		for item in result:
			if item["blug"] == 0 or item['staaddress'] == u'长虹':
				continue
			if u'测试' in item['staname']:
				continue 
			provinceId = int(item['province'])
			provinceItem = province[provinceId]
			result1 = self.charger.findPortPerStation(item['staid'],item['blug'])	#每个充电站的结果
			if result1 == "false":
				continue
			ports = self.charger.filiterLostStatusData(result1)				# 过滤每个结果，得到异常的数据
			length = len(ports)
			if length != 0:
				data = {'province':int(item['province']), 'statname':item['staname'], 'address':item['staaddress'], 'data':ports}
				allUnormalData.append(data)
			#获取订单任务
			orderData = self.getOrdersInfo(item['staname'], timeArrays[0], timeArrays[1])
			all_proj.setdefault(item['staname'],{'order': orderData,'proj_name': item['staname'],'province':provinceItem['name'], 'all_port': item['blug'], 'lost_port':length, 'error_port':0, 'new_lost_port': 0})
		return self.saveData(allUnormalData, oldFileData, filename, all_proj)
	
	#获取规定时间内的该电站的所有订单
	#响应的charge_status字段解释：没有 -- 正在充电 0 - 用户拔掉插头 1 - 电量充满  2 - 电流过载 3 -订单结束 4- 正常结束 
	def getOrdersInfo(self, staName, timeS, timeE):
		ret = self.charger.getStationOrder(staName, timeS, timeE)
		if ret == 'false' or ret == '':
			return
		allOrderCount = 0
		less30Count = 0
		less30User = ''
		charingCount = 0
		#分析结果
		for item in ret:
			if staName != item['staName']:
				continue
			allOrderCount = allOrderCount + 1
			try:
				stat = item['charge_status']
				if item['chargeDuration'] <= 30:
					less30Count = less30Count + 1
					less30User = less30User + item['userName'] + '-' + str(item['charge_status']) + u'状态，'+ str(item['chargeDuration']) + u'分钟'+ '\n'
			except KeyError:
				charingCount = charingCount + 1
		data = {'all': allOrderCount, 'less30': less30Count, 'charging': charingCount, 'less30user': less30User}
		return data
	


	#获取所有数据  正常和异常
	def getAllDataFromNetwork(self, filename, code):
		print("---start---")
		province = self.charger.getProvinceInfo()
		result = self.charger.listAllProvinceData()
		if result == "false" or province == 'false':
			return "false"
		allData = []
		all_proj = {}
		all_gate = []
		for item in result:
			if item["blug"] == 0:
				continue
			provinceId = int(item['province'])
			provinceItem = province[provinceId]
			result_port = self.charger.findPortPerStation(item['staid'],item['blug'])	#每个充电站的充电端口结果
			result_gateway = self.charger.findGateway(item['staid'])
			if result_port == "false" or result_gateway == 'false':
				continue
			length = len(result_port)
			if length != 0:
				data = {'province':int(item['province']), 'statname':item['staname'], 'address':item['staaddress'], 'data':result_port}
				allData.append(data)
			all_proj.setdefault(item['staname'], {'proj_name': item['staname'],'province':provinceItem['name'], 'all_port': item['blug'], 'lost_port':length, 'new_lost_port': 0})
			length_gateway = len(result_gateway)
			if length_gateway != 0:
				data1 = {'statname':item['staname'], 'data':result_gateway}
				all_gate.append(data1)
		#self.saveGatewayData(all_gate)	
		return self.saveData(allData, None, filename, all_proj)
			
	#save网关
	def saveGatewayData(self, data):
		book = Workbook()	
		sheet0 = book.create_sheet(u'网关表格')
		sheet0.append(['电站名', '网关号'])
		for item in data:
			staname = item['statname']
			for iitem in item['data']:
				savedata = [staname, iitem['devimeiStr']]
				sheet0.append(savedata)
		book.save('gateway.xls')


	def saveAllStaConfData(self, allData, filename):
		book = Workbook()
                sheet0 = book.create_sheet(u'异常详细表格', 0)
                sheet0.append(['充电庄异常信息表','',datetime.datetime.now().strftime('%Y-%m-%d')])
                sheet0.append([''])
                sheet0.append(['充电站名称','欠载阈值电流','过载阈值电流','满电量阈值电流','空载判定时间', '过载判定时间'])
		for item in allData:
			data = [ item['staname'],  item['under_load_electric'], item['over_load_electric'],item['full_electric'], item['no_load_time'], item['over_load_time']]
			sheet0.append(data)
			print '1'
		book.save(filename)

	#保存数据到xls
	# 类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
	def saveData(self, netWorkData, oldFileData, filename, all_proj):
		total_new_count = 0
		book = Workbook()
		sheet0 = book.create_sheet(u'异常详细表格', 0)
		sheet1 = book.create_sheet(u'异常总览表', 1) 
		sheet0.append(['充电庄异常信息表','',datetime.datetime.now().strftime('%Y-%m-%d')])
		sheet0.append([''])
		sheet0.append(['地区','充电站名称','端口编号','状态','详细地址','备注'])
		sheet1.append(['各地区异常数量汇总表'])
		sheet1.append([''])
		sheet1.append(['地区', '项目名称', '所有端口', '失联数量', '故障数量', '新增异常数量', '总订单', '小于30分钟订单', '正在充电订单', '小于30分钟订单用户状态'])
		for parent_item in netWorkData:
			data = parent_item['data']
			current_proj = all_proj[parent_item['statname']]
			province_name = current_proj['province']
			new_lost_count = 0
			eData = []
			errorNumber = 0
			for child_item in data:
				status = child_item['pgstatus']
				str_statu = '失联'
				if status == 2:
					str_statu = '故障'
					errorNumber = errorNumber + 1
				portNum = child_item['pgnum']
				if oldFileData is None or 'add' !=  oldFileData.startCompare(portNum):
					eData = [province_name, parent_item['statname'],portNum,str_statu, parent_item['address']]
				else:
					new_lost_count = 1 + new_lost_count
					eData = [province_name, parent_item['statname'],portNum,str_statu, parent_item['address'], 'add']
				self.saveDataToExcel(sheet0, eData)
			current_proj['new_lost_port'] = new_lost_count
			current_proj['error_port'] = errorNumber
			current_proj['lost_port'] = current_proj['lost_port'] - errorNumber
		totalLostCount = 0	#所有失联类型端口
		totalErrorCount = 0	#所有故障类型端口
		totalPortCount = 0	 #所有端口数量
		totalNewPort = 0	#新增失联+故障
		for item in all_proj:
			value = all_proj[item]
			order = value['order']
			sheet1.append([value['province'], item ,value['all_port'],value['lost_port'], value['error_port'], value['new_lost_port'], order['all'], order['less30'], order['charging'],order['less30user']])
			totalLostCount = totalLostCount + value['lost_port']
			totalPortCount = totalPortCount + value['all_port']
			totalErrorCount = totalErrorCount + value['error_port']
			totalNewPort = totalNewPort + value['new_lost_port']
		sheet1.append(['合计', '', totalPortCount, totalLostCount, totalErrorCount, totalNewPort])
		book.save(filename)
		print('---end---')
		return 'true'

	def saveDataToExcel(self, sheet, data):
		sheet.append(data)


	def form_xls_add_device(self, filepath, code):
		ret = self.charger.login(code)
		value = self.init_baimingdan(filepath)
		return  value

	def init_baimingdan(self, filepath):
		fail = 0
		sucess = 0
		failed_flag = 0
		if os.path.isfile(filepath):
			wb = xlrd.open_workbook(filepath)
			sheets = wb.sheets()	#获取所有表格的名字
			sheet0 = sheets[0]			#获取第一个表格
			rows = sheet0.nrows
			key0 = sheet0.cell(0,0).value
			key1 = sheet0.cell(0,1).value
			key2 = sheet0.cell(0,2).value
			key3 = sheet0.cell(0,3).value
			rows = rows - 1
			
			for row in range(rows):
				data = {}
				data['rows'] = []
				row_data = {}
				row_data[key0] = sheet0.cell(row+1,0).value
				row_data[key1] = sheet0.cell(row+1,1).value
				row_data[key2] = sheet0.cell(row+1,2).value
				row_data[key3] = sheet0.cell(row+1,3).value
				if row_data[key0] == '' or row_data[key1] == '' or row_data[key1] == '' or row_data[key3] == '':
					continue
				data['rows'].append(row_data)
				value = self.charger.add_device_data(data)
				if value != 1:
					sucess = sucess + 1
					failed_flag = 1
					print data['rows']
					print value['message'].encode('utf-8')
				else:
					fail = fail + 1
		else:
			return '批量文件错误'
		if failed_flag  == 0:
			return '添加白名单成功'
		else:
			str1 = '部分设备数据添加失败，请查看日志\n'
			return str1
					


	def main(self, oldFile, newFile, code, time):
		ret = self.charger.login(code)
		if ret != 'true':
			return ret
		com = None
		if os.path.isfile(oldFile):
			exec_utils = CompareEXCEL(oldFile)		#获得excel对比工具
		else:
			exec_utils = None
		return self.getLostDataFromNetwork(exec_utils,newFile, code, time)

	def getAllProjectConfig(self, fileName, code):
		ret = self.charger.login(code)
		if ret != 'true':
			return ret
		return self.getStationConfig(fileName)

if __name__ == '__main__':
	print('bash use rules:')
	print('parameter 1~2: oldExceFile, newExcelFile')
#	main(sys.argv[1], sys.argv[2])
#	getAllProject(sys.argv[1])
	charg = EasyCharger('test')	
	test = ChargerManager(charg)
	test.form_xls_add_device(None)

